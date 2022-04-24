from google.cloud import language
from google.cloud.language import enums, types
import pandas as pd
from openpyxl import load_workbook

# name of the excel file
myExcel = '20-25-april2.xlsx'

# name of the sheet
sheet_name = 'Barometre2304'
# Specify the path of the text that needs to be analyzed, open the file and then read it!
text = open('C:/Users/hamra/PycharmProjects/gcp_nlp_api/What is private equity, and why is it killing everything you love.txt')
inputText = text.read()

# ======================================= SENTIMENT ANALYSIS ======================================================


def analyze_text_sentiment(text):
    client = language.LanguageServiceClient()
    document = types.Document(
        content=text,
        type=enums.Document.Type.PLAIN_TEXT)

    response = client.analyze_sentiment(document=document)

    # making  a list to put all the elements of sentiment analysis
    sentence_text = [sentence.text.content for sentence in response.sentences]
    sentence_sentiment_score = [sentence.sentiment.score for sentence in response.sentences]
    sentence_sentiment_magnitude = [sentence.sentiment.magnitude for sentence in response.sentences]
    # making a data frame with pandas in order to structure the data.
    sentiment_analysis = pd.DataFrame({
        'Sentences': sentence_text,
        'Score' : sentence_sentiment_score,
        'Magnitude': sentence_sentiment_magnitude
    })
    # sentiment_analysis.to_csv(myExcel)
    # Writing the data frame to an excel file
    writer = pd.ExcelWriter(myExcel, engine='openpyxl', mode='a')
    writer.book = load_workbook(myExcel)
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    reader = pd.read_excel(myExcel, sheet_name=sheet_name)
    sentiment_analysis.to_excel(writer, sheet_name=sheet_name, index=False, startrow=len(reader)+4)
    writer.save()

    # Overall Sentiment
    sentiment = response.document_sentiment
    overall_sentiment = [sentiment.score]
    overall_magnitude = [sentiment.magnitude]
    results = pd.DataFrame({
        'Overall Sentiment': overall_sentiment,
        'Overall Magnitude': overall_magnitude
    })
    # Create a pandas Excel writer using openpyxl as the engine
    writer = pd.ExcelWriter(myExcel, engine='openpyxl', mode='a')
    # open an existing workbook
    writer.book = load_workbook(myExcel)
    # copy existing sheets
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    # read the existing file
    reader = pd.read_excel(myExcel, sheet_name=sheet_name)
    # convert the dataframe to an openpyxl excel object
    results.to_excel(writer, sheet_name=sheet_name, index=False, header=True, startrow=len(reader)+1)
    # close the pandas excel writer and output the excel file
    writer.save()


analyze_text_sentiment(inputText)


# ======================================= CONTENT CLASSIFICATION =======================================

def classify_text(text):
    client = language.LanguageServiceClient()
    document = types.Document(
        content=text,
        type=enums.Document.Type.PLAIN_TEXT)

    response = client.classify_text(document=document)

    category_name = [category.name for category in response.categories]
    category_confidence = [category.confidence for category in response.categories]

    text_classification = pd.DataFrame({
        'Category Name': category_name,
        'Category_confidence': category_confidence
    })
    # with open("Sentiment_Analysis_PE2104.csv",mode='a') as myfile:
    #    myfile.write("\n"*2)
    # text_classification.to_excel("Sentiment_Analysis_PE2104.csv", mode='a')

    writer = pd.ExcelWriter(myExcel, engine='openpyxl', mode='a')
    writer.book = load_workbook(myExcel)
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    reader = pd.read_excel(myExcel, sheet_name=sheet_name)
    text_classification.to_excel(writer, index=False, header=True, sheet_name=sheet_name, startrow=len(reader)+2)
    writer.save()


classify_text(inputText)

# ======================================= ENTITY ANALYSIS (also includes entity sentiment) =======================================


def analyze_text_entities(text):
    client = language.LanguageServiceClient()
    document = types.Document(
        content=text,
        type=enums.Document.Type.PLAIN_TEXT)

    response = client.analyze_entity_sentiment(document=document)
    entity_name = [entity.name for entity in response.entities]
    entity_type = [enums.Entity.Type(entity.type).name for entity in response.entities]
    entity_salience = [entity.salience for entity in response.entities]
    entity_wikipedia = [entity.metadata.get('wikipedia_url', '-') for entity in response.entities]
    entity_mid = [entity.metadata.get('mid', '-') for entity in response.entities]
    entity_sentiment_score = [entity.sentiment.score for entity in response.entities]
    entity_sentiment_magnitude = [entity.sentiment.magnitude for entity in response.entities]

    entity_analysis = pd.DataFrame({
        'Entity Name': entity_name,
        'Entity Type': entity_type,
        'Entity Salience': entity_salience,
        'Wikipedia': entity_wikipedia,
        'Mid': entity_mid,
        'Entity Score': entity_sentiment_score,
        'Entity Magnitude': entity_sentiment_magnitude
    })
    writer = pd.ExcelWriter(myExcel, engine='openpyxl', mode='a')
    writer.book = load_workbook(myExcel)
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    reader = pd.read_excel(myExcel, sheet_name=sheet_name)
    entity_analysis.to_excel(writer, index=False, header=True, sheet_name=sheet_name, startrow=len(reader)+2)
    writer.save()


analyze_text_entities(inputText)
# ======================================= SYNTAX ANALYSIS =======================================


def analyze_text_syntax(text):
    client = language.LanguageServiceClient()
    document = types.Document(
        content=text,
        type=enums.Document.Type.PLAIN_TEXT)

    response = client.analyze_syntax(document=document)
    num_of_senteces = [len(response.sentences)]
    num_of_tokens = [len(response.tokens)]

    text_info = pd.DataFrame({
        'Number of Senteces': num_of_senteces,
        'Number of Tokens': num_of_tokens
    })

    writer = pd.ExcelWriter(myExcel, engine='openpyxl', mode='a')
    writer.book = load_workbook(myExcel)
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    reader = pd.read_excel(myExcel, sheet_name=sheet_name)
    text_info.to_excel(writer, sheet_name=sheet_name, index=False, header=True, startrow=len(reader)+2)
    writer.save()

    tag_name = []
    tag_name_content = []
    # voice = []
    tense = []
    lemma = []
    for token in response.tokens:
        part_of_speech = token.part_of_speech
        tag_name.append(enums.PartOfSpeech.Tag(part_of_speech.tag).name)
        tag_name_content.append(token.text.content)
        # voice.append(enums.PartOfSpeech.Voice(part_of_speech.voice).name)
        tense.append(enums.PartOfSpeech.Tense(part_of_speech.tense).name)
        lemma.append(token.lemma)

    text_sysntax = pd.DataFrame({
        'Part of Speech': tag_name,
        'Content': tag_name_content,
        'Tense': tense,
        # 'Voice': voice,
        'Lemma': lemma
    })
    # Create a pandas Excel writer using openpyxl as the engine
    writer = pd.ExcelWriter(myExcel, engine='openpyxl', mode='a')
    # open an existing workbook
    writer.book = load_workbook(myExcel)
    # copy existing sheets
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    # read the existing file
    reader = pd.read_excel(myExcel, sheet_name=sheet_name)
    # convert the dataframe to an openpyxl excel object
    text_sysntax.to_excel(writer, sheet_name=sheet_name, index=False, header=True, startrow=len(reader) + 1)
    # close the pandas excel writer and output the excel file
    writer.save()


analyze_text_syntax(inputText)
